import io
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image, ImageOps, ImageEnhance

import fitz  # PyMuPDF
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ----------------------------
# CONFIG / REGEX
# ----------------------------
TECO_FULL_RE = re.compile(r"TECO[\s\-\:_]*([0-9]{4,6})", re.IGNORECASE)
DIGITS_RE = re.compile(r"\d{1,4}")

DEFAULT_ZOOM = 2.5

# Measurement search window relative to TECO bbox (in pixels of rendered image)
DEFAULT_MEAS_XPAD = 180
DEFAULT_MEAS_Y_UP = 520
DEFAULT_MEAS_Y_DOWN = 70

# OCR thresholds
DEFAULT_MEAS_MIN_CONF = 45.0  # below -> mark as "da verificare"

YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BOLD_FONT = Font(bold=True)

# ----------------------------
# DATA STRUCTURES
# ----------------------------
@dataclass
class Detection:
    page: int
    rotation: int
    code: str
    teco_bbox: Tuple[int, int, int, int]  # x,y,w,h on rotated image
    meters_guess: Optional[int]
    meters_conf: float
    status: str
    teco_crop_png: bytes
    meas_crop_png: bytes


# ----------------------------
# IMAGE / OCR HELPERS
# ----------------------------
def render_pdf_page(pdf_bytes: bytes, pno: int, zoom: float) -> Image.Image:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[pno]
    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
    img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
    return img


def pil_to_png_bytes(img: Image.Image) -> bytes:
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def enhance_for_printed_text(img: Image.Image) -> Image.Image:
    g = ImageOps.grayscale(img)
    g = ImageEnhance.Contrast(g).enhance(2.2)
    g = ImageEnhance.Sharpness(g).enhance(1.8)
    return g


def enhance_for_pencil_digits(img: Image.Image) -> Image.Image:
    g = ImageOps.grayscale(img)
    g = ImageEnhance.Contrast(g).enhance(3.0)
    g = ImageEnhance.Sharpness(g).enhance(2.2)
    # soft threshold (pencil is light)
    bw = g.point(lambda x: 0 if x < 185 else 255)
    return bw


def ocr_words_with_boxes(img: Image.Image, config: str) -> pd.DataFrame:
    """
    Returns DataFrame with columns: text, conf, left, top, width, height, line_num, block_num, par_num.
    """
    data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT, config=config)
    df = pd.DataFrame(data)
    # Normalize
    df["text"] = df["text"].fillna("").astype(str)
    df["conf"] = pd.to_numeric(df["conf"], errors="coerce").fillna(-1)
    df = df[df["text"].str.strip() != ""].copy()
    return df


def detect_teco_boxes_on_image(img: Image.Image) -> List[Tuple[str, Tuple[int, int, int, int], float]]:
    """
    Detect TECO codes with bbox from OCR tokens.
    Returns list of (code, bbox(x,y,w,h), conf)
    """
    cfg = "--psm 6"
    df = ocr_words_with_boxes(enhance_for_printed_text(img), config=cfg)

    out = []
    texts = df["text"].tolist()
    for i, row in df.iterrows():
        t = str(row["text"]).strip()
        t_norm = re.sub(r"\s+", "", t.upper())

        # Case 1: TECO18278 in one token
        m = TECO_FULL_RE.search(t_norm)
        if m:
            code = m.group(1)
            bbox = (int(row["left"]), int(row["top"]), int(row["width"]), int(row["height"]))
            out.append((code, bbox, float(row["conf"])))
            continue

        # Case 2: token 'TECO' and digits in next tokens (same line)
        if t_norm == "TECO":
            # try up to next 3 tokens
            for j in range(1, 4):
                if i + j not in df.index:
                    continue
                r2 = df.loc[i + j]
                t2 = str(r2["text"]).strip()
                t2_norm = re.sub(r"\D", "", t2)
                if len(t2_norm) in (4, 5, 6):
                    code = t2_norm
                    # bbox around number token (better reference for measure)
                    bbox = (int(r2["left"]), int(r2["top"]), int(r2["width"]), int(r2["height"]))
                    conf = float(r2["conf"])
                    out.append((code, bbox, conf))
                    break

    # Deduplicate: same code very close -> keep highest conf
    dedup = {}
    for code, bbox, conf in out:
        key = (code, bbox[0] // 20, bbox[1] // 20)
        if key not in dedup or conf > dedup[key][2]:
            dedup[key] = (code, bbox, conf)
    return list(dedup.values())


def guess_measure_near(img: Image.Image, bbox: Tuple[int, int, int, int],
                       xpad: int, y_up: int, y_down: int) -> Tuple[Optional[int], float, Image.Image]:
    """
    Search for digits in a window above the TECO bbox.
    Returns (meters_guess, conf_max, crop_img).
    """
    x, y, w, h = bbox
    cx = x + w // 2
    x1 = max(0, cx - xpad)
    x2 = min(img.size[0], cx + xpad)
    y1 = max(0, y - y_up)
    y2 = max(0, y - y_down)

    if y2 <= y1 or x2 <= x1:
        crop = img.crop((max(0, x), max(0, y - 150), min(img.size[0], x + 350), y + 50))
        return None, -1.0, crop

    crop = img.crop((x1, y1, x2, y2))
    crop_enh = enhance_for_pencil_digits(crop)

    cfg = "--psm 6 -c tessedit_char_whitelist=0123456789"
    data = pytesseract.image_to_data(crop_enh, output_type=pytesseract.Output.DICT, config=cfg)
    best_val = None
    best_conf = -1.0

    for txt, conf in zip(data.get("text", []), data.get("conf", [])):
        if not txt:
            continue
        t = str(txt).strip()
        if not t.isdigit():
            continue
        v = int(t)
        # plausible meters
        if v <= 0 or v > 5000:
            continue
        try:
            cf = float(conf)
        except Exception:
            cf = 0.0
        if cf > best_conf:
            best_conf = cf
            best_val = v

    return best_val, best_conf, crop


def choose_best_rotation(img: Image.Image) -> Tuple[int, Image.Image, List[Tuple[str, Tuple[int,int,int,int], float]]]:
    """
    Try 0/90/180/270 and choose the rotation with the most TECO detections.
    Returns (rotation, rotated_img, detections)
    """
    best = (0, img, [])
    for rot in (0, 90, 180, 270):
        imr = img.rotate(rot, expand=True)
        det = detect_teco_boxes_on_image(imr)
        if len(det) > len(best[2]):
            best = (rot, imr, det)
    return best


# ----------------------------
# EXCEL HELPERS
# ----------------------------
def find_header_columns(ws) -> Dict[str, int]:
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        key = str(v).strip().upper()
        headers[key] = c
    return headers


def read_known_codes_from_excel(xlsx_bytes: bytes) -> Tuple[pd.DataFrame, List[str]]:
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    headers = find_header_columns(ws)

    # Minimal: must have CODICE; MT is where we'll write
    # If not present, still read the table for codes
    col_code = headers.get("CODICE", None)
    if col_code is None:
        raise ValueError("Nell'Excel non trovo la colonna 'CODICE' in riga 1.")

    # Load rows into DataFrame
    rows = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, col_code).value
        if code is None or str(code).strip() == "":
            continue
        code_s = re.sub(r"\D", "", str(code))
        if code_s:
            rows.append({"CODICE": code_s})
    df = pd.DataFrame(rows).drop_duplicates()
    return df, df["CODICE"].tolist()


def update_excel_with_totals(xlsx_bytes: bytes,
                            totals: Dict[str, int],
                            raw: List[Detection],
                            illegibili: List[Detection]) -> bytes:
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    headers = find_header_columns(ws)

    if "CODICE" not in headers or "MT" not in headers:
        raise ValueError("Nell'Excel non trovo le colonne richieste: 'CODICE' e 'MT' (header riga 1).")

    col_code = headers["CODICE"]
    col_mt = headers["MT"]

    # Map existing codes to row
    existing = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_code).value
        if v is None:
            continue
        code_s = re.sub(r"\D", "", str(v))
        if code_s:
            existing[code_s] = r

    # Write totals; add unknown codes in new rows (highlight)
    last_row = ws.max_row
    for code, mt in sorted(totals.items(), key=lambda x: int(x[0])):
        if code in existing:
            ws.cell(row=existing[code], column=col_mt).value = int(mt)
        else:
            last_row += 1
            ws.cell(last_row, col_code).value = int(code)
            ws.cell(last_row, col_mt).value = int(mt)
            for c in range(1, ws.max_column + 1):
                ws.cell(last_row, c).fill = YELLOW_FILL
            ws.cell(last_row, col_code).font = BOLD_FONT
            ws.cell(last_row, col_mt).font = BOLD_FONT

    # Remove / recreate detail sheets
    for name in ("TWINPALL_RAW", "TWINPALL_SUM", "ILLEGIBILI"):
        if name in wb.sheetnames:
            del wb[name]

    ws_raw = wb.create_sheet("TWINPALL_RAW")
    ws_sum = wb.create_sheet("TWINPALL_SUM")
    ws_ill = wb.create_sheet("ILLEGIBILI")

    # RAW
    ws_raw.append(["PAGINA", "ROT", "CODICE", "METRI", "CONF", "STATUS"])
    for d in raw:
        ws_raw.append([d.page, d.rotation, d.code, d.meters_guess, round(d.meters_conf, 1), d.status])

    # SUM
    ws_sum.append(["CODICE", "MT_TOTALE"])
    for code, mt in sorted(totals.items(), key=lambda x: int(x[0])):
        ws_sum.append([int(code), int(mt)])

    # ILLEGIBILI
    ws_ill.append(["PAGINA", "ROT", "CODICE", "STATUS"])
    for d in illegibili:
        ws_ill.append([d.page, d.rotation, d.code, d.status])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ----------------------------
# STREAMLIT UI
# ----------------------------
st.set_page_config(page_title="Twinpall • Lista Cavi → Excel MT", layout="wide")

st.title("Twinpall • Estrazione metrature TECO e aggiornamento Excel")

st.markdown(
    """
Questa app:
- legge il PDF (lista cavi **Twinpall**) e cerca i codici **TECOxxxxx**
- prova a leggere la misura (metri) scritta a matita **vicino al TECO**
- ti fa **revisionare** le misure dubbie
- scrive i totali nella colonna **MT** dell’Excel (aggiungendo righe nuove evidenziate se un codice non esiste)
"""
)

with st.sidebar:
    st.header("File")
    pdf_file = st.file_uploader("Carica PDF lista cavi (Twinpall)", type=["pdf"])
    xlsx_file = st.file_uploader("Carica Excel magazzino (con colonne CODICE, MT)", type=["xlsx"])

    st.divider()
    st.header("Parametri OCR")
    zoom = st.slider("Zoom render PDF", 1.5, 4.0, float(DEFAULT_ZOOM), 0.1)

    xpad = st.slider("Finestra misura: X padding", 60, 350, DEFAULT_MEAS_XPAD, 10)
    y_up = st.slider("Finestra misura: quanto sopra al TECO", 150, 900, DEFAULT_MEAS_Y_UP, 10)
    y_down = st.slider("Finestra misura: margine sotto (esclude TECO)", 20, 200, DEFAULT_MEAS_Y_DOWN, 5)

    min_conf = st.slider("Soglia confidenza misura", 0.0, 95.0, float(DEFAULT_MEAS_MIN_CONF), 1.0)

    st.divider()
    st.header("Pagine")
    page_mode = st.radio("Selezione pagine", ["Tutte", "Range"], horizontal=True)
    p_from, p_to = 1, 1
    if page_mode == "Range":
        p_from = st.number_input("Da pagina (1-based)", min_value=1, value=1, step=1)
        p_to = st.number_input("A pagina (1-based)", min_value=1, value=1, step=1)

if pdf_file is None or xlsx_file is None:
    st.info("Carica PDF + Excel per iniziare.")
    st.stop()

pdf_bytes = pdf_file.getvalue()
xlsx_bytes = xlsx_file.getvalue()

# Read known codes from Excel
try:
    codes_df, known_codes = read_known_codes_from_excel(xlsx_bytes)
except Exception as e:
    st.error(f"Errore Excel: {e}")
    st.stop()

doc = fitz.open(stream=pdf_bytes, filetype="pdf")
num_pages = len(doc)

if page_mode == "Range":
    p_from = int(max(1, min(p_from, num_pages)))
    p_to = int(max(p_from, min(p_to, num_pages)))
    page_indices = list(range(p_from - 1, p_to))
else:
    page_indices = list(range(num_pages))

st.caption(f"Pagine PDF: {num_pages} • Pagine selezionate: {len(page_indices)} • Codici in Excel: {len(known_codes)}")

if "detections" not in st.session_state:
    st.session_state["detections"] = []  # List[Detection]

if "run_done" not in st.session_state:
    st.session_state["run_done"] = False

colA, colB = st.columns([1, 1])
with colA:
    run = st.button("1) Avvia estrazione (OCR)", type="primary")
with colB:
    reset = st.button("Reset risultati")

if reset:
    st.session_state["detections"] = []
    st.session_state["run_done"] = False
    st.rerun()

if run:
    detections: List[Detection] = []
    prog = st.progress(0, text="Elaborazione…")
    for idx, pno in enumerate(page_indices):
        img = render_pdf_page(pdf_bytes, pno, zoom=zoom)
        rot, imr, tecos = choose_best_rotation(img)

        # Filter TECO codes: keep those that are in Excel, but also keep unknown (to add later)
        for code, bbox, conf in tecos:
            # Make a TECO crop for preview
            x, y, w, h = bbox
            teco_crop = imr.crop((max(0, x - 40), max(0, y - 25), min(imr.size[0], x + w + 40), min(imr.size[1], y + h + 25)))
            teco_crop_png = pil_to_png_bytes(teco_crop)

            meters, mconf, mcrop = guess_measure_near(imr, bbox, xpad=xpad, y_up=y_up, y_down=y_down)
            meas_crop_png = pil_to_png_bytes(mcrop)

            status = "OK"
            if meters is None:
                status = "MISURA_ILLEGIBILE"
            elif mconf < min_conf:
                status = f"DA_VERIFICARE (conf={mconf:.1f})"

            detections.append(Detection(
                page=pno + 1,
                rotation=rot,
                code=code,
                teco_bbox=bbox,
                meters_guess=meters,
                meters_conf=float(mconf),
                status=status,
                teco_crop_png=teco_crop_png,
                meas_crop_png=meas_crop_png
            ))

        prog.progress((idx + 1) / len(page_indices), text=f"Elaborazione pagina {pno+1}/{num_pages}…")

    st.session_state["detections"] = detections
    st.session_state["run_done"] = True
    prog.empty()

if not st.session_state["run_done"]:
    st.stop()

detections: List[Detection] = st.session_state["detections"]

if not detections:
    st.warning("Nessun TECO rilevato nelle pagine selezionate. Prova ad aumentare lo zoom o cambiare range pagine.")
    st.stop()

# Build editable table
table_rows = []
for i, d in enumerate(detections):
    in_excel = "SI" if d.code in known_codes else "NO"
    table_rows.append({
        "ID": i,
        "PAGINA": d.page,
        "CODICE": d.code,
        "IN_EXCEL": in_excel,
        "METRI_LETTI": d.meters_guess,
        "CONF": round(d.meters_conf, 1),
        "STATUS": d.status,
        "USA_METRI": d.meters_guess if d.status.startswith("OK") else d.meters_guess,
        "NOTE": ""
    })

df = pd.DataFrame(table_rows)

st.subheader("2) Revisione misure (consigliato)")
st.markdown(
    """
- **USA_METRI** è la colonna che verrà usata per la somma finale.  
- Se una misura è sbagliata o illeggibile, correggi **USA_METRI** (oppure lasciala vuota) e scrivi una **NOTE**.
"""
)

edited = st.data_editor(
    df,
    use_container_width=True,
    num_rows="fixed",
    column_config={
        "ID": st.column_config.NumberColumn(disabled=True),
        "PAGINA": st.column_config.NumberColumn(disabled=True),
        "CODICE": st.column_config.TextColumn(disabled=True),
        "IN_EXCEL": st.column_config.TextColumn(disabled=True),
        "METRI_LETTI": st.column_config.NumberColumn(disabled=True),
        "CONF": st.column_config.NumberColumn(disabled=True),
        "STATUS": st.column_config.TextColumn(disabled=True),
        "USA_METRI": st.column_config.NumberColumn(help="Metri finali usati nella somma (puoi modificarli)"),
        "NOTE": st.column_config.TextColumn(help="Annotazioni (es. 'illeggibile', 'da verificare')"),
    },
    hide_index=True
)

with st.expander("Anteprima crop (TECO / MISURA) per ogni riga"):
    for i, d in enumerate(detections):
        r = edited.loc[edited["ID"] == i].iloc[0]
        st.markdown(f"**ID {i} — Pag. {d.page} — TECO{d.code} — STATUS: {d.status} — USA_METRI: {r['USA_METRI']}**")
        c1, c2 = st.columns(2)
        with c1:
            st.image(d.teco_crop_png, caption="Crop TECO (codice)", use_container_width=True)
        with c2:
            st.image(d.meas_crop_png, caption="Crop misura (a matita)", use_container_width=True)
        st.divider()

# Compute totals from edited table
totals: Dict[str, int] = {}
illegibili: List[Detection] = []
raw: List[Detection] = []

for i, d in enumerate(detections):
    raw.append(d)
    row = edited.loc[edited["ID"] == i].iloc[0]
    val = row["USA_METRI"]
    if pd.isna(val) or val is None:
        # mark as illegible / not used
        illegibili.append(d)
        continue
    try:
        v = int(val)
    except Exception:
        illegibili.append(d)
        continue
    if v <= 0:
        illegibili.append(d)
        continue
    totals[d.code] = totals.get(d.code, 0) + v

st.subheader("3) Totali per codice (da USA_METRI)")
sum_df = pd.DataFrame(
    [{"CODICE": k, "MT_TOTALE": v, "IN_EXCEL": ("SI" if k in known_codes else "NO")} for k, v in sorted(totals.items(), key=lambda x: int(x[0]))]
)
st.dataframe(sum_df, use_container_width=True, hide_index=True)

st.caption(f"Righe illeggibili/non conteggiate: {len(illegibili)} (saranno nel foglio 'ILLEGIBILI').")

gen = st.button("4) Genera Excel aggiornato", type="primary")

if gen:
    try:
        out_bytes = update_excel_with_totals(xlsx_bytes, totals, raw, illegibili)
        st.success("Excel generato.")
        st.download_button(
            "Scarica Excel compilato",
            data=out_bytes,
            file_name="CAVI_COMMESSE_TWINPALL_COMPILATO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.info("Controlla il foglio 'ILLEGIBILI' e i crop nell’expander per verifiche rapide.")
    except Exception as e:
        st.error(f"Errore generazione Excel: {e}")
